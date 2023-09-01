# -*- coding: utf-8 -*-
"""
Created on Tue Aug 29 16:23:11 2023

@author: Terry.Li
"""

import openpyxl
import numpy as np
from scipy.signal import find_peaks
import matplotlib.pyplot as plt
import pandas as pd
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import tkinter as tk
from tkinter import filedialog

def smooth2_data(input_file_name, same_num,interval):

    # 读取Excel表单
    input_file = input_file_name
       
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    
    # 获取波长和counts数据
    wavelength_column = sheet['X']
    
    counts_column = sheet['Y']
    
    NUM = len(counts_column)
    
    Originalx = list(range(1, len(counts_column)))
    
    Originalwavelengths = [cell.value for cell in wavelength_column[1:]]
    
    Originalcounts = [cell.value for cell in counts_column[1:]]
    
    
    # 进行局部拟合并找到peak值的像素位置
    Originalsmoothed_counts = np.convolve(Originalcounts, np.ones(same_num)/same_num, mode='same')  # 平滑处理
    
   
    Originalpeak_indices, _ = find_peaks(Originalsmoothed_counts)
    
    # 找到最大的peak值对应的像素位置
    Originalmax_peak_index = Originalpeak_indices[np.argmax(Originalsmoothed_counts[Originalpeak_indices])] +1
    
    Originalmax_peak_counts = Originalx[Originalmax_peak_index]
    
     
    x = list(range(Originalmax_peak_counts-interval, Originalmax_peak_counts+interval))
    
  
    
    wavelengths = [cell.value for cell in wavelength_column[1:]]
    
    counts = [cell.value for cell in counts_column[Originalmax_peak_counts-interval:Originalmax_peak_counts+interval]]
    
    
    def fitfun(x, a, b, c):
        return a*(x-b)**2 + c
        #return a*x**2 + b*x+ c
    
    def fund(x,a,b,c):
        return a*np.sin(2*np.pi/12*x+b)+c
    
    popt, pcov = curve_fit(fitfun, x, counts, p0=[3,2,-16])
    
    
    a, b, c = popt
    
    x_model = np.linspace(min(x), max(x),interval*2)
        
    y_model = fitfun(x_model, a, b, c)
    
    peak_indices, _ = find_peaks(y_model)
    
       
    # 找到最大的peak值对应的像素位置
    max_peak_index = peak_indices[np.argmax(y_model[peak_indices])] + Originalmax_peak_counts-interval+1
    
    max_peak_wavelength = wavelengths[max_peak_index-1]
    
    
    
    #计算R2值
    r2_value = r2_score(counts, y_model)

    print("R2值:", r2_value)
        
    # plot1 = plt.plot(x, counts, 'r*', label='original values')
    # plot2 = plt.plot(x, [fitfun(i, *popt) for i in x], 'b', label='curvefit values')
    
    
    # plt.axvline(x=max_peak_index, color='b', linestyle='--', label=f'max_peak_index = {max_peak_index}')
    
    # plt.axvline(x=max_peak_counts, color='r', linestyle='--', label=f'max_peak_index = {max_peak_counts}')

     
    #向excel中写入表头
    sheet['e1'] = 'Max_index_wavelength'
        
    #向excel中写入对应的value
    sheet.cell(row=2, column=5).value = max_peak_index
    sheet.cell(row=3, column=5).value = max_peak_wavelength
    workbook.save(input_file)
     
        
    return y_model,Originalcounts,max_peak_index,max_peak_wavelength,Originalmax_peak_counts


def run_smooth():
    input_file = file_path_entry.get()
    same_num = int(same_num_entry.get())
    interval = int(interval_entry.get())
    
    smoothed_counts, Originalcounts, max_peak_counts, max_peak_wavelength, Originalmax_peak_counts = smooth2_data(input_file, same_num, interval)

    df = pd.DataFrame({'wavelength_Data': smoothed_counts})
    df.to_csv('output1.csv', index=False)

    x = list(range(Originalmax_peak_counts - interval, Originalmax_peak_counts + interval))

    plt.plot(Originalcounts, label="Original")
    plt.plot(x, smoothed_counts, label="Smoothed")

    plt.axvline(x=max_peak_counts, color='gray', linestyle='--', label=f'Max_counts = {max_peak_counts} Max_wavelength = {max_peak_wavelength} ')

    plt.xlabel("Pixel")
    plt.ylabel("Smoothed Pixels")
    plt.legend()
    plt.show()

    print(max_peak_counts)
    print(f"Max Peak Wavelength: {max_peak_wavelength}")

# 创建GUI窗口
root = tk.Tk()
root.title("Smooth2 Data GUI")

# 添加文件选择按钮和运行按钮
file_path_label = tk.Label(root, text="Input File Path:")
file_path_label.pack()

file_path_entry = tk.Entry(root)
file_path_entry.pack()

select_file_button = tk.Button(root, text="Select File", command=lambda: file_path_entry.insert(0, filedialog.askopenfilename()))
select_file_button.pack()

same_num_label = tk.Label(root, text="Same Num:")
same_num_label.pack()

same_num_entry = tk.Entry(root)
same_num_entry.pack()

interval_label = tk.Label(root, text="Interval:")
interval_label.pack()

interval_entry = tk.Entry(root)
interval_entry.pack()

run_button = tk.Button(root, text="Run Smooth2", command=run_smooth)
run_button.pack()

root.mainloop()
